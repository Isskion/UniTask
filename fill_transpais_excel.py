import openpyxl

excel_path = r'C:\Users\daniel.delamo\Downloads\transpais_plantilla_integracion (2).xlsx'
wb = openpyxl.load_workbook(excel_path)

# Let's inspect TransicionesTSP sheet
ws_tsp = wb['TransicionesTSP']

# Headers
headers = [ws_tsp.cell(1, c).value for c in range(1, ws_tsp.max_column + 1)]
print("Original headers TransicionesTSP:", headers)

# Ensure 'operacion' is the first column header if not present
if 'operacion' not in headers:
    ws_tsp.insert_cols(1)
    ws_tsp.cell(1, 1, 'operacion')
    headers = ['operacion'] + headers

print("Updated headers TransicionesTSP:", headers)

# Clear existing sample rows (rows 2 onwards)
while ws_tsp.max_row > 1:
    ws_tsp.delete_rows(2)

# Definición de transiciones para Distribución (operación = 'Distribución')
# Formato: [operacion, entidad, estadoOrigen, estadoDestino, requiereFoto, requiereFirma, requiereMotivo, validarGeocerca, idEstadoParadaVisita, triggerCascada, disparadoPor, nota, activo]
transiciones = [
    # --- PEDIDO ---
    ['Distribución', 'Pedido', 'INGRESADO', 'ERROR-REQUIERE AJUSTE', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Falta de datos o geocodificación', 'TRUE'],
    ['Distribución', 'Pedido', 'INGRESADO', 'GRABADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Validación automática OK', 'TRUE'],
    ['Distribución', 'Pedido', 'ERROR-REQUIERE AJUSTE', 'GRABADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Ajuste manual de pedido', 'TRUE'],
    ['Distribución', 'Pedido', 'GRABADO', 'CONFIRMADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Customer Service confirma', 'TRUE'],
    ['Distribución', 'Pedido', 'CONFIRMADO', 'PROGRAMAR REPARTO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, 'Orden = 102', None, 'Planificador asigna ruta de reparto', 'TRUE'],
    ['Distribución', 'Pedido', 'CONFIRMADO', 'PROGRAMAR DIRECTO REMITENTE DESTINO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, 'Orden = 102', None, 'Planificador asigna servicio directo', 'TRUE'],
    ['Distribución', 'Pedido', 'CONFIRMADO', 'PROGRAMAR RECOLECCIÓN', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, 'Orden = 102', None, 'Planificador asigna recogida', 'TRUE'],
    ['Distribución', 'Pedido', 'PROGRAMAR REPARTO', 'ENTREGADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, 'Parada', 'POD firmado en destino', 'TRUE'],
    ['Distribución', 'Pedido', 'PROGRAMAR REPARTO', 'ENTREGA PARCIAL', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, 'Parada', 'Entrega parcial en destinatario', 'TRUE'],
    ['Distribución', 'Pedido', 'PROGRAMAR REPARTO', 'NO ENTREGADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, 'Parada', 'Rechazo/ausente en destinatario', 'TRUE'],
    ['Distribución', 'Pedido', 'PROGRAMAR RECOLECCIÓN', 'RECOLECTADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, 'Parada', 'Recogida completada en remitente', 'TRUE'],
    ['Distribución', 'Pedido', 'PROGRAMAR RECOLECCIÓN', 'NO RECOLECTADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, 'Parada', 'Recogida fallida', 'TRUE'],
    ['Distribución', 'Pedido', 'PROGRAMAR RECOLECCIÓN', 'RECOLECTADO PARCIAL', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, 'Parada', 'Recogida parcial', 'TRUE'],
    ['Distribución', 'Pedido', 'ENTREGADO', 'LIQUIDADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, 'Viaje', 'Preliquidación enviada a BC', 'TRUE'],
    ['Distribución', 'Pedido', 'ENTREGA PARCIAL', 'LIQUIDADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, 'Viaje', 'Preliquidación enviada a BC', 'TRUE'],
    ['Distribución', 'Pedido', 'RECOLECTADO', 'LIQUIDADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, 'Viaje', 'Preliquidación enviada a BC', 'TRUE'],
    ['Distribución', 'Pedido', 'RECOLECTADO PARCIAL', 'LIQUIDADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, 'Viaje', 'Preliquidación enviada a BC', 'TRUE'],

    # --- ORDEN ---
    ['Distribución', 'Orden', 'PENDIENTE', 'PLANIFICADA', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, 'Parada = 203, Viaje = 105', None, 'Ruta creada y asignada', 'TRUE'],
    ['Distribución', 'Orden', 'PLANIFICADA', 'EN TRÁNSITO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Inicio de viaje / salida a reparto', 'TRUE'],
    ['Distribución', 'Orden', 'EN TRÁNSITO', 'FINALIZADA', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Parada de entrega ejecutada', 'TRUE'],
    ['Distribución', 'Orden', 'EN TRÁNSITO', 'RECOLECTADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Parada de recogida ejecutada', 'TRUE'],

    # --- RUTA ---
    ['Distribución', 'Ruta', 'CREADA', 'EN RUTA', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Viaje activado', 'TRUE'],
    ['Distribución', 'Ruta', 'EN RUTA', 'FINALIZADA', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Todas las paradas finalizadas', 'TRUE'],

    # --- VIAJE ---
    ['Distribución', 'Viaje', 'INACTIVO', 'ASIGNADO / PENDIENTE', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Vehículo y conductor asignados', 'TRUE'],
    ['Distribución', 'Viaje', 'ASIGNADO / PENDIENTE', 'CONFIRMADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Asignación aceptada y recursos validados', 'TRUE'],
    ['Distribución', 'Viaje', 'ASIGNADO / PENDIENTE', 'RECHAZADO', 'FALSE', 'FALSE', 'TRUE', 'FALSE', None, None, None, 'Transportista rechaza asignación', 'TRUE'],
    ['Distribución', 'Viaje', 'RECHAZADO', 'ASIGNADO / PENDIENTE', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Reasignación de transporte', 'TRUE'],
    ['Distribución', 'Viaje', 'CONFIRMADO', 'ACTIVO / EN EJECUCIÓN', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, 'Ruta = 201', None, 'Salida a reparto / seguimiento activo', 'TRUE'],
    ['Distribución', 'Viaje', 'ACTIVO / EN EJECUCIÓN', 'FINALIZADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, 'Ruta = 401', None, 'Paradas completadas', 'TRUE'],
    ['Distribución', 'Viaje', 'FINALIZADO', 'RENDIDO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Rendición de almacén / Interfaz 9 WMS', 'TRUE'],
    ['Distribución', 'Viaje', 'RENDIDO', 'LIQUIDABLE', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, None, None, 'Incidencias aprobadas por Tráfico', 'TRUE'],
    ['Distribución', 'Viaje', 'LIQUIDABLE', 'LIQUIDADO', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, 'Pedido = 502', None, 'Preliquidaciones confirmadas BC', 'TRUE'],

    # --- PARADA ---
    ['Distribución', 'Parada', 'PENDIENTE', 'EN VIAJE', 'FALSE', 'FALSE', 'FALSE', 'FALSE', None, 'Orden = 202', None, 'Inicio de tramo a parada', 'TRUE'],
    ['Distribución', 'Parada', 'EN VIAJE', 'VISITADO / EN GEOCERCA', 'FALSE', 'FALSE', 'FALSE', 'TRUE', 1, None, None, 'GPS automático — Arribo en geocerca', 'TRUE'],
    ['Distribución', 'Parada', 'EN VIAJE', 'VISITADO / EN GEOCERCA', 'FALSE', 'FALSE', 'FALSE', 'TRUE', 2, None, None, 'GPS automático — Arribo a tiempo', 'TRUE'],
    ['Distribución', 'Parada', 'EN VIAJE', 'VISITADO / EN GEOCERCA', 'FALSE', 'FALSE', 'FALSE', 'TRUE', 3, None, None, 'GPS automático — Fuera de horario', 'TRUE'],
    ['Distribución', 'Parada', 'EN VIAJE', 'VISITADO / EN GEOCERCA', 'FALSE', 'FALSE', 'FALSE', 'TRUE', 4, None, None, 'GPS automático — Fuera de orden', 'TRUE'],

    ['Distribución', 'Parada', 'EN VIAJE', 'CARGADO', 'FALSE', 'TRUE', 'FALSE', 'FALSE', None, None, None, 'Carga en almacén OK', 'TRUE'],
    ['Distribución', 'Parada', 'EN VIAJE', 'CARGADO PARCIAL', 'FALSE', 'TRUE', 'TRUE', 'FALSE', None, None, None, 'Falta de stock / bulto dañado', 'TRUE'],
    ['Distribución', 'Parada', 'EN VIAJE', 'NO CARGADO', 'TRUE', 'FALSE', 'TRUE', 'FALSE', None, None, None, 'Incidencia grave en carga', 'TRUE'],

    ['Distribución', 'Parada', 'VISITADO / EN GEOCERCA', 'CARGADO', 'FALSE', 'TRUE', 'FALSE', 'FALSE', None, None, None, 'Carga en muelle OK', 'TRUE'],
    ['Distribución', 'Parada', 'VISITADO / EN GEOCERCA', 'CARGADO PARCIAL', 'FALSE', 'TRUE', 'TRUE', 'FALSE', None, None, None, 'Carga incompleta en muelle', 'TRUE'],
    ['Distribución', 'Parada', 'VISITADO / EN GEOCERCA', 'NO CARGADO', 'TRUE', 'FALSE', 'TRUE', 'FALSE', None, None, None, 'No cargado en muelle', 'TRUE'],

    ['Distribución', 'Parada', 'EN VIAJE', 'ENTREGADO', 'TRUE', 'TRUE', 'FALSE', 'FALSE', None, 'Orden = 306, Pedido = 303', None, 'POD (Firma + Foto). Interfaz 8 CM Logistics', 'TRUE'],
    ['Distribución', 'Parada', 'EN VIAJE', 'ENTREGA PARCIAL', 'TRUE', 'TRUE', 'TRUE', 'FALSE', None, 'Orden = 306, Pedido = 304', None, 'Genera Parada Rendición Almacén', 'TRUE'],
    ['Distribución', 'Parada', 'EN VIAJE', 'NO ENTREGADO', 'TRUE', 'FALSE', 'TRUE', 'FALSE', None, 'Pedido = 305', None, 'Genera Parada Rendición Almacén', 'TRUE'],

    ['Distribución', 'Parada', 'VISITADO / EN GEOCERCA', 'ENTREGADO', 'TRUE', 'TRUE', 'FALSE', 'FALSE', None, 'Orden = 306, Pedido = 303', None, 'POD (Firma + Foto). Interfaz 8 CM Logistics', 'TRUE'],
    ['Distribución', 'Parada', 'VISITADO / EN GEOCERCA', 'ENTREGA PARCIAL', 'TRUE', 'TRUE', 'TRUE', 'FALSE', None, 'Orden = 306, Pedido = 304', None, 'Genera Parada Rendición Almacén', 'TRUE'],
    ['Distribución', 'Parada', 'VISITADO / EN GEOCERCA', 'NO ENTREGADO', 'TRUE', 'FALSE', 'TRUE', 'FALSE', None, 'Pedido = 305', None, 'Genera Parada Rendición Almacén', 'TRUE'],

    ['Distribución', 'Parada', 'EN VIAJE', 'RECOLECTADO EN DEVOLUCIÓN', 'FALSE', 'TRUE', 'FALSE', 'FALSE', None, 'Orden = 400, Pedido = 400', None, 'Recogida completada en cliente', 'TRUE'],
    ['Distribución', 'Parada', 'EN VIAJE', 'NO RECOLECTADO', 'TRUE', 'FALSE', 'TRUE', 'FALSE', None, 'Pedido = 404', None, 'Recogida fallida', 'TRUE'],
    ['Distribución', 'Parada', 'EN VIAJE', 'RECOLECTADO PARCIAL', 'FALSE', 'TRUE', 'TRUE', 'FALSE', None, 'Orden = 400, Pedido = 405', None, 'Recogida parcial', 'TRUE'],

    ['Distribución', 'Parada', 'VISITADO / EN GEOCERCA', 'RECOLECTADO EN DEVOLUCIÓN', 'FALSE', 'TRUE', 'FALSE', 'FALSE', None, 'Orden = 400, Pedido = 400', None, 'Recogida completada en cliente', 'TRUE'],
    ['Distribución', 'Parada', 'VISITADO / EN GEOCERCA', 'NO RECOLECTADO', 'TRUE', 'FALSE', 'TRUE', 'FALSE', None, 'Pedido = 404', None, 'Recogida fallida', 'TRUE'],
    ['Distribución', 'Parada', 'VISITADO / EN GEOCERCA', 'RECOLECTADO PARCIAL', 'FALSE', 'TRUE', 'TRUE', 'FALSE', None, 'Orden = 400, Pedido = 405', None, 'Recogida parcial', 'TRUE'],
]

for row in transiciones:
    ws_tsp.append(row)

# Also let's update 'Transiciones' sheet if present
if 'Transiciones' in wb.sheetnames:
    ws_t = wb['Transiciones']
    # If header doesn't have operacion, insert operacion header
    t_headers = [ws_t.cell(1, c).value for c in range(1, ws_t.max_column + 1)]
    if 'operacion' not in t_headers:
        ws_t.insert_cols(1)
        ws_t.cell(1, 1, 'operacion')

wb.save(excel_path)
print(f"Successfully populated {len(transiciones)} Distribución transitions into {excel_path}")
